"""
從 Excel 匯入訂單資料
"""
import openpyxl
import uuid
from datetime import datetime
from database import SessionLocal, Order, Product, BOM, ComponentSchedule, Inventory, MoldCalculation, Completion, init_db
import math

def parse_date(value):
    """解析日期格式"""
    if not value:
        return None
    
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    
    # 處理字串格式
    try:
        value_str = str(value).strip()
        if not value_str or value_str == 'None':
            return None
        
        # 嘗試不同的日期格式
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y%m%d', '%m/%d/%Y']:
            try:
                dt = datetime.strptime(value_str, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        # 如果都失敗，返回當前日期
        return datetime.now().strftime('%Y-%m-%d')
    except:
        return datetime.now().strftime('%Y-%m-%d')

def import_orders_from_excel(file_path):
    """從 Excel 匯入訂單"""
    
    # 讀取 Excel
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 初始化資料庫
    init_db()
    db = SessionLocal()
    
    try:
        imported_count = 0
        skipped_count = 0
        updated_count = 0
        
        # 讀取表頭
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers[header.strip()] = col
        
        print(f"找到的欄位: {list(headers.keys())}")
        
        # 第一遍：收集所有訂單資料，檢查重複
        order_totals = {}  # {(order_number, product_code): total_quantity}
        order_data = {}    # {(order_number, product_code): {due_date, order_date, ...}}
        seen_duplicates = set()  # 記錄重複的 (訂單單號, 訂單序, 訂單數量) 組合
        
        for row in range(2, ws.max_row + 1):
            try:
                order_number = ws.cell(row, headers.get('訂單單號', 0)).value if '訂單單號' in headers else None
                product_code = ws.cell(row, headers.get('品號', 0)).value if '品號' in headers else None
                quantity = ws.cell(row, headers.get('訂單數量', 0)).value if '訂單數量' in headers else None
                order_sequence = ws.cell(row, headers.get('訂單序', 0)).value if '訂單序' in headers else None
                
                if not order_number or not product_code:
                    continue
                
                order_number_str = str(order_number).strip()
                product_code_str = str(product_code).strip()
                order_sequence_str = str(order_sequence).strip() if order_sequence else ""
                
                try:
                    quantity_int = int(float(quantity)) if quantity else 0
                except (ValueError, TypeError):
                    quantity_int = 0
                
                if quantity_int <= 0:
                    continue
                
                # 重複檢查：訂單單號 + 訂單序 + 訂單數量
                duplicate_key = (order_number_str, order_sequence_str, quantity_int)
                if duplicate_key in seen_duplicates:
                    print(f"  跳過重複資料：訂單 {order_number_str}, 序號 {order_sequence_str}, 數量 {quantity_int}")
                    skipped_count += 1
                    continue
                
                seen_duplicates.add(duplicate_key)
                key = (order_number_str, product_code_str)
                
                # 累加數量
                if key in order_totals:
                    order_totals[key] += quantity_int
                else:
                    order_totals[key] = quantity_int
                    # 保存第一筆的訂單資訊
                    order_data[key] = {
                        'due_date': ws.cell(row, headers.get('預定到達日', 0)).value if '預定到達日' in headers else None,
                        'order_date': ws.cell(row, headers.get('接單日期', 0)).value if '接單日期' in headers else None,
                        'customer_id': ws.cell(row, headers.get('客戶編號', 0)).value if '客戶編號' in headers else None,
                        'order_sequence': order_sequence_str,
                    }
            except Exception as e:
                continue
        
        print(f"\n收集到 {len(order_totals)} 個唯一的訂單+品號組合")
        
        # 第二遍：處理每個訂單
        for (order_number_str, product_code_str), quantity_int in order_totals.items():
            try:
                data = order_data[(order_number_str, product_code_str)]
                
                # 解析日期
                due_date_str = parse_date(data['due_date'])
                order_date_str = parse_date(data['order_date'])
                customer_id_str = str(data['customer_id']).strip() if data['customer_id'] else None
                order_sequence_str = str(data['order_sequence']).strip() if data['order_sequence'] else None
                
                # 檢查訂單+品號組合是否已存在
                existing_order = db.query(Order).filter(
                    Order.order_number == order_number_str,
                    Order.product_code == product_code_str
                ).first()
                
                if existing_order:
                    # 重新匯入：刪除舊的 Product 和 ComponentSchedule，重建資料
                    print(f"  訂單 {order_number_str} 品號 {product_code_str} 已存在，清理舊資料並重建（總量: {quantity_int}）...")
                    
                    # 刪除舊的 Product 記錄
                    db.query(Product).filter(Product.order_id == existing_order.id).delete()
                    
                    # 刪除舊的 ComponentSchedule 記錄
                    db.query(ComponentSchedule).filter(ComponentSchedule.order_id == existing_order.id).delete()
                    
                    # 查詢庫存，計算未交數量
                    order_inventory = db.query(Inventory).filter(Inventory.product_code == product_code_str).first()
                    order_current_inventory = order_inventory.quantity if order_inventory else 0
                    order_undelivered_qty = max(0, quantity_int - order_current_inventory)
                    
                    # 更新訂單資訊
                    existing_order.due_date = due_date_str or existing_order.due_date
                    existing_order.order_date = order_date_str
                    existing_order.customer_id = customer_id_str
                    existing_order.customer_name = customer_id_str or existing_order.customer_name
                    existing_order.order_sequence = order_sequence_str
                    existing_order.quantity = quantity_int  # 使用累加後的總量
                    existing_order.undelivered_quantity = order_undelivered_qty  # 訂單量 - 庫存
                    existing_order.updated_at = datetime.utcnow()
                    
                    print(f"    更新訂單 {order_number_str}: 訂單量={quantity_int}, 庫存={order_current_inventory}, 未交數量={order_undelivered_qty}")
                    
                    order_id = existing_order.id
                    updated_count += 1
                else:
                    # 查詢庫存，計算未交數量
                    new_order_inventory = db.query(Inventory).filter(Inventory.product_code == product_code_str).first()
                    new_order_current_inventory = new_order_inventory.quantity if new_order_inventory else 0
                    new_order_undelivered_qty = max(0, quantity_int - new_order_current_inventory)
                    
                    # 創建新訂單
                    order_id = str(uuid.uuid4())
                    new_order = Order(
                        id=order_id,
                        order_number=order_number_str,
                        customer_name=customer_id_str or "未知客戶",
                        customer_id=customer_id_str,
                        product_code=product_code_str,
                        quantity=quantity_int,
                        undelivered_quantity=new_order_undelivered_qty,  # 訂單量 - 庫存
                        due_date=due_date_str or datetime.now().strftime('%Y-%m-%d'),
                        order_date=order_date_str,
                        order_sequence=order_sequence_str,
                        priority=3,
                        status="PENDING"
                    )
                    db.add(new_order)
                    db.flush()
                    imported_count += 1
                    
                    print(f"    新建訂單 {order_number_str}: 訂單量={quantity_int}, 庫存={new_order_current_inventory}, 未交數量={new_order_undelivered_qty}")
                
                # 統一創建產品記錄（無論新建或更新）
                # 創建產品記錄（0階成品）
                # 查詢庫存，計算未交數量 = 訂單量 - 庫存
                inventory = db.query(Inventory).filter(Inventory.product_code == product_code_str).first()
                current_inventory = inventory.quantity if inventory else 0
                undelivered_qty = max(0, quantity_int - current_inventory)  # 未交數量不能為負
                
                product = Product(
                    id=str(uuid.uuid4()),
                    order_id=order_id,
                    product_code=product_code_str,
                    quantity=quantity_int,
                    undelivered_quantity=undelivered_qty,  # 初始未交 = 訂單總量 - 庫存
                    product_type='finished'  # 0階成品
                )
                db.add(product)
                db.flush()
                
                print(f"    成品 {product_code_str}: 訂單量={quantity_int}, 庫存={current_inventory}, 未交數量={undelivered_qty}")
                
                # 自動拆解成子件（BOM展開）
                bom_items = db.query(BOM).filter(BOM.product_code == product_code_str).all()
                
                if bom_items:
                    # 分離模具和非模具子件
                    mold_items = [item for item in bom_items if item.component_code.startswith('6')]
                    non_mold_items = [item for item in bom_items if not item.component_code.startswith('6')]
                    
                    # 先處理非模具子件
                    non_mold_quantities = {}  # 記錄非模具子件的數量
                    for bom_item in non_mold_items:
                        # 查詢子件庫存
                        component_inventory = db.query(Inventory).filter(
                            Inventory.product_code == bom_item.component_code
                        ).first()
                        component_stock = component_inventory.quantity if component_inventory else 0
                        
                        required_quantity = quantity_int  # 子件總量
                        undelivered_quantity = max(0, quantity_int - component_stock)  # 總量 - 庫存量
                        non_mold_quantities[bom_item.component_code] = required_quantity
                        
                        print(f"      子件 {bom_item.component_code}: 需求量={required_quantity}, 庫存={component_stock}, 未交數量={undelivered_quantity}")
                        
                        # 創建非模具子件產品記錄
                        component_product = Product(
                            id=str(uuid.uuid4()),
                            order_id=order_id,
                            product_code=bom_item.component_code,
                            quantity=required_quantity,
                            undelivered_quantity=undelivered_quantity,
                            product_type='component'  # 1階子件
                        )
                        db.add(component_product)
                        db.flush()
                    
                    # 再處理模具子件（6開頭）
                    for bom_item in mold_items:
                        # 查詢模具庫存
                        component_inventory = db.query(Inventory).filter(
                            Inventory.product_code == bom_item.component_code
                        ).first()
                        component_stock = component_inventory.quantity if component_inventory else 0
                        
                        # 模具數量計算：找同個order_id中1開頭子件的未交數量
                        # 先找到該訂單中所有1開頭的子件
                        one_prefix_products = db.query(Product).filter(
                            Product.order_id == order_id,
                            Product.product_code.like('1%'),
                            Product.product_type == 'component'
                        ).all()
                        
                        # 使用第一個1開頭子件的未交數量作為基準（如果沒有則用成品未交數量）
                        if one_prefix_products:
                            base_undelivered_qty = one_prefix_products[0].undelivered_quantity
                        else:
                            # 如果沒有1開頭子件，使用成品未交數量
                            base_undelivered_qty = undelivered_qty
                        
                        # 查詢模具的穴數，優先查詢 BOM 表
                        mold_calc = db.query(MoldCalculation).filter(
                            (MoldCalculation.component_code == bom_item.component_code) |
                            (MoldCalculation.mold_code == bom_item.component_code)
                        ).first()
                        
                        cavity_count = mold_calc.cavity_count if mold_calc and mold_calc.cavity_count else 1
                        
                        # 模具數量 = ceil(1開頭子件未交數量 / 模具穴數)
                        required_quantity = math.ceil(base_undelivered_qty / cavity_count) if base_undelivered_qty > 0 else 0
                        undelivered_quantity = max(0, required_quantity - component_stock)
                        
                        print(f"      模具 {bom_item.component_code}: 1開頭子件未交量={base_undelivered_qty}, 穴數={cavity_count}, 計算: ceil({base_undelivered_qty}/{cavity_count})={required_quantity}, 庫存={component_stock}, 模具未交數量={undelivered_quantity}")
                        
                        # 創建模具子件產品記錄
                        component_product = Product(
                            id=str(uuid.uuid4()),
                            order_id=order_id,
                            product_code=bom_item.component_code,
                            quantity=required_quantity,
                            undelivered_quantity=undelivered_quantity,
                            product_type='component'  # 1階子件
                        )
                        db.add(component_product)
                        db.flush()
                    
                    # 統一處理所有子件的排程記錄
                    for bom_item in bom_items:
                        # 獲取已創建的子件記錄來確定未交數量
                        component_product = db.query(Product).filter(
                            Product.order_id == order_id,
                            Product.product_code == bom_item.component_code,
                            Product.product_type == 'component'
                        ).first()
                        
                        if not component_product:
                            continue
                        
                        # 判斷初始狀態
                        if bom_item.component_code.startswith('6'):
                            initial_status = "模具"  # 6開頭是模具,不需排程
                        elif component_product.undelivered_quantity == 0:
                            initial_status = "無法進行排程"  # 數量為0不排程
                        else:
                            # 檢查是否有完整的 mold_calculations 資料
                            # 必須有機台、穴數>0、成型時間>0
                            has_complete_mold_calc = db.query(MoldCalculation).filter(
                                MoldCalculation.component_code == bom_item.component_code,
                                MoldCalculation.machine_id.isnot(None),
                                MoldCalculation.cavity_count.isnot(None),
                                MoldCalculation.cavity_count > 0,
                                MoldCalculation.avg_molding_time_sec.isnot(None),
                                MoldCalculation.avg_molding_time_sec > 0
                            ).first() is not None
                            
                            if has_complete_mold_calc:
                                initial_status = "未排程"  # 有完整資料可排程
                            else:
                                initial_status = "無法進行排程"  # 沒有完整的模具計算資料
                        
                        # 創建或更新元件排程記錄（使用未交數量）
                        existing_schedule = db.query(ComponentSchedule).filter(
                            ComponentSchedule.order_id == order_id,
                            ComponentSchedule.component_code == bom_item.component_code
                        ).first()
                        
                        if existing_schedule:
                            # 已存在，累加數量
                            existing_schedule.quantity += component_product.undelivered_quantity
                            print(f"    累加排程記錄 {bom_item.component_code} 數量: +{component_product.undelivered_quantity} → {existing_schedule.quantity}")
                        else:
                            # 不存在，創建新的排程記錄
                            component_schedule = ComponentSchedule(
                                id=str(uuid.uuid4()),
                                order_id=order_id,
                                component_code=bom_item.component_code,
                                quantity=component_product.undelivered_quantity,  # 使用未交數量
                                status=initial_status
                            )
                            db.add(component_schedule)
                
                # 記錄處理數量
                print(f"  ✓ 已處理訂單 {order_number_str}（成品+{len(bom_items) if bom_items else 0}個子件）")
                
                # 每100筆提交一次
                if (imported_count + updated_count) % 100 == 0:
                    db.commit()
                    print(f"已處理 {imported_count + updated_count} 筆...")
                
            except Exception as e:
                import traceback
                print(f"處理第 {row} 行時出錯: {e}")
                print(f"  訂單號: {order_number}, 品號: {product_code}")
                traceback.print_exc()
                skipped_count += 1
                continue
        
        # 最後提交
        db.commit()
        
        print("\n" + "="*60)
        print(f"✓ 匯入完成！")
        print(f"  新增: {imported_count} 筆")
        print(f"  更新: {updated_count} 筆")
        print(f"  跳過: {skipped_count} 筆")
        print("="*60)
        
        # 對齊子件數量（確保模具回次和子件數量正確）
        print("\n正在對齊子件數量...")
        try:
            alignment_count = 0
            all_orders = db.query(Order).all()
            
            for order in all_orders:
                finished_products = db.query(Product).filter(
                    Product.order_id == order.id,
                    Product.product_type == 'finished'
                ).all()
                
                component_products = db.query(Product).filter(
                    Product.order_id == order.id,
                    Product.product_type == 'component'
                ).all()
                
                for finished in finished_products:
                    # 找到該訂單中1開頭的子件作為基準
                    one_prefix_components = [comp for comp in component_products if comp.product_code.startswith('1')]
                    base_undelivered_qty = one_prefix_components[0].undelivered_quantity if one_prefix_components else finished.undelivered_quantity
                    
                    for comp in component_products:
                        if comp.product_code.startswith('6'):
                            # 模具：使用1開頭子件未交數量重新計算
                            mold_calc = db.query(MoldCalculation).filter(
                                (MoldCalculation.mold_code == comp.product_code) |
                                (MoldCalculation.component_code == comp.product_code)
                            ).first()
                            cavity_count = mold_calc.cavity_count if mold_calc and mold_calc.cavity_count else 1
                            expected_undelivered = math.ceil(base_undelivered_qty / cavity_count) if base_undelivered_qty > 0 else 0
                            
                            if comp.undelivered_quantity != expected_undelivered:
                                comp.undelivered_quantity = expected_undelivered
                                alignment_count += 1
                        else:
                            # 子件：等於成品未交數量（quantity 保持不變）
                            if comp.undelivered_quantity != finished.undelivered_quantity:
                                comp.undelivered_quantity = finished.undelivered_quantity
                                alignment_count += 1
            
            db.commit()
            print(f"✓ 子件數量對齊完成：調整了 {alignment_count} 筆記錄")
        except Exception as e:
            print(f"⚠️  子件對齊時出錯: {e}")
        
        # 同步 Product → ComponentSchedule
        print("\n正在同步到排程記錄...")
        try:
            sync_count = 0
            all_schedules = db.query(ComponentSchedule).all()
            
            for schedule in all_schedules:
                product = db.query(Product).filter(
                    Product.order_id == schedule.order_id,
                    Product.product_code == schedule.component_code
                ).first()
                
                if product and schedule.quantity != product.undelivered_quantity:
                    schedule.quantity = product.undelivered_quantity
                    sync_count += 1
            
            db.commit()
            print(f"✓ 排程記錄同步完成：更新了 {sync_count} 筆記錄")
        except Exception as e:
            print(f"⚠️  排程同步時出錯: {e}")
        
        # 重新應用 Completion 記錄
        print("\n正在重新應用完工記錄...")
        try:
            from main import update_undelivered_quantity
            
            # 查詢所有完工記錄，按時間順序應用
            all_completions = db.query(Completion).order_by(Completion.completion_date, Completion.id).all()
            
            if all_completions:
                print(f"找到 {len(all_completions)} 筆完工記錄")
                for completion in all_completions:
                    update_undelivered_quantity(db, completion.finished_item_no, completion.completed_qty)
                db.commit()
                print(f"✓ 完工記錄應用完成：處理了 {len(all_completions)} 筆記錄")
            else:
                print("✓ 無完工記錄需要應用")
        except Exception as e:
            print(f"⚠️  應用完工記錄時出錯: {e}")
            import traceback
            traceback.print_exc()
        
        # 自動觸發模具計算
        calc_warnings = []
        try:
            print("\n正在計算模具需求...")
            from mold_calc import calculate_and_save
            import traceback
            calc_result = calculate_and_save(silent=True, save_excel=False)
            if calc_result.get('success'):
                print(f"✓ 模具計算完成：{calc_result.get('count', 0)} 筆資料已更新")
                calc_warnings = calc_result.get('warnings', [])
                if calc_warnings:
                    print("\n⚠️  警示訊息:")
                    for warning in calc_warnings:
                        print(f"  {warning}")
            else:
                print(f"⚠️  模具計算返回失敗: {calc_result}")
        except Exception as e:
            print(f"⚠️  模具計算失敗（不影響訂單匯入）: {e}")
            traceback.print_exc()
        
        return {
            "success": True,
            "imported": imported_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "warnings": calc_warnings
        }
        
    except Exception as e:
        db.rollback()
        print(f"❌ 匯入過程中發生錯誤: {e}")
        raise
    finally:
        db.close()

if __name__ == "__main__":
    # 測試匯入
    import_orders_from_excel("raw_data/未交訂單EX.xlsx")
